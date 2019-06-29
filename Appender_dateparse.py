
# coding: utf-8

# In[11]:


import pandas as pd
import numpy as np
import os

FileList = os.listdir()

if not os.path.exists("Appending"):
    os.mkdir("Appending") 


# In[12]:


xlsx_files = list(filter(lambda x: x.endswith("xlsx"), FileList ))

#убрать результирующий файл отсюда
xlsx_df = {xlsx_file : pd.read_excel(xlsx_file,comment= "#", encoding = "UTF-8",parse_dates=['Дата'], dayfirst=True ) for xlsx_file in xlsx_files }

Total = pd.concat(xlsx_df)


# In[13]:


from datetime import datetime
timestamp= datetime.now().strftime("%Y%m%d-%H%M%S")
Total.to_excel("Appending\\Appended_"+timestamp +".xlsx", merge_cells=False)


# In[14]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""

